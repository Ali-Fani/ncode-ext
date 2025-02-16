import os
import re
import time
from pathlib import Path
from tqdm import tqdm
from colorama import init, Fore, Style
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Initialize colorama
init()

def print_ascii_art():
    """Print a cool ASCII art banner"""
    ascii_art = """
    ███████╗██╗██╗     ███████╗    ███████╗ ██████╗ █████╗ ███╗   ██╗
    ██╔════╝██║██║     ██╔════╝    ██╔════╝██╔════╝██╔══██╗████╗  ██║
    █████╗  ██║██║     █████╗      ███████╗██║     ███████║██╔██╗ ██║
    ██╔══╝  ██║██║     ██╔══╝      ╚════██║██║     ██╔══██║██║╚██╗██║
    ██║     ██║███████╗███████╗    ███████║╚██████╗██║  ██║██║ ╚████║
    ╚═╝     ╚═╝╚══════╝╚══════╝    ╚══════╝ ╚═════╝╚═╝  ╚═╝╚═╝  ╚═══╝
    """
    print(Fore.CYAN + ascii_art + Style.RESET_ALL)

def scan_files():
    """
    Scan files in the current directory for national numbers and create an Excel report.
    This version uses openpyxl directly to build the Excel file, removing the need for pandas.
    """
    print_ascii_art()
    time.sleep(0.5)  # Pause for dramatic effect

    print(Fore.CYAN + "=== Starting File Scanner ===" + Style.RESET_ALL)

    # Configuration: precompile regex pattern for matching 9 or 10 digit national numbers
    pattern = r'(?<![0-9\u06F0-\u06F9])((?:[0-9\u06F0-\u06F9]{10})|(?:[0-9\u06F0-\u06F9]{9}))(?![0-9\u06F0-\u06F9])'
    regex = re.compile(pattern)

    root_dir = os.getcwd()
    print(Fore.CYAN + f"📂 Scanning directory: {root_dir}" + Style.RESET_ALL)

    matching_results = []

    # Count total files using os.walk to avoid building a huge list in memory
    total_files = sum(len(files) for _, _, files in os.walk(root_dir))
    print(Fore.YELLOW + f"📑 Scanning {total_files} file(s) for matching national codes..." + Style.RESET_ALL)

    pbar = tqdm(total=total_files, desc="🔍 Scanning Files", bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.GREEN, Style.RESET_ALL))
    for dirpath, _, filenames in os.walk(root_dir):
        for filename in filenames:
            pbar.update(1)
            matches = regex.findall(filename)
            for match in matches:
                national_number = match
                if len(national_number) == 9:
                    national_number = "0" + national_number
                matching_results.append({
                    'NationalNumber': national_number,
                    'FilePath': os.path.join(dirpath, filename)
                })
                # Optionally comment out the line below for performance when many matches are found
                # print(Fore.GREEN + f"✅ Match: '{filename}' -> National Number: {national_number}" + Style.RESET_ALL)
    pbar.close()

    print(Fore.CYAN + f"✨ Finished scanning. Found {len(matching_results)} matching file(s)." + Style.RESET_ALL)

    if not matching_results:
        print(Fore.RED + "❌ No matches found. Exiting..." + Style.RESET_ALL)
        return

    # Create a list for unique national numbers (keeping the first occurrence)
    unique_results = []
    seen = set()
    for entry in matching_results:
        if entry['NationalNumber'] not in seen:
            seen.add(entry['NationalNumber'])
            unique_results.append(entry)

    print(Fore.GREEN + f"📊 Found {len(unique_results)} unique national number(s)." + Style.RESET_ALL)

    # Create an Excel workbook using openpyxl
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Matches"
    ws_all.append(["NationalNumber", "FilePath"])  # Header

    for entry in matching_results:
        ws_all.append([entry['NationalNumber'], entry['FilePath']])

    # Format "All Matches" sheet as a table
    all_rows = ws_all.max_row
    data_range_all = f"A1:B{all_rows}"
    table_all = Table(displayName="Table_All_Matches", ref=data_range_all)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=True,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table_all.tableStyleInfo = style
    ws_all.add_table(table_all)
    ws_all.column_dimensions['A'].width = 20  # National Number column
    ws_all.column_dimensions['B'].width = 100  # File Path column

    # Format first column as text to preserve leading zeros
    for row in ws_all.iter_rows(min_row=2, max_row=all_rows, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = '@'

    # Create a new sheet for Unique Numbers
    ws_unique = wb.create_sheet(title="Unique Numbers")
    ws_unique.append(["NationalNumber", "FilePath"])  # Header

    for entry in unique_results:
        ws_unique.append([entry['NationalNumber'], entry['FilePath']])

    unique_rows = ws_unique.max_row
    data_range_unique = f"A1:B{unique_rows}"
    table_unique = Table(displayName="Table_Unique_Numbers", ref=data_range_unique)
    table_unique.tableStyleInfo = style
    ws_unique.add_table(table_unique)
    ws_unique.column_dimensions['A'].width = 20
    ws_unique.column_dimensions['B'].width = 100

    for row in ws_unique.iter_rows(min_row=2, max_row=unique_rows, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = '@'

    output_path = os.path.join(root_dir, "matching_files.xlsx")
    wb.save(output_path)

    print(Fore.CYAN + """
╔════════════════════════════════════╗
║ ✨ Excel file created successfully ✨ ║
╚════════════════════════════════════╝
""" + Style.RESET_ALL)

if __name__ == "__main__":
    scan_files()
