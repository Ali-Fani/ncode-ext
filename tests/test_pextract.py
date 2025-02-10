import unittest
import os
import shutil
import pandas as pd
from pathlib import Path
from pextract import scan_files

class TestFileScanner(unittest.TestCase):
    def setUp(self):
        """Create a temporary test directory with sample files"""
        # Create test directory
        self.test_dir = Path("test_files")
        self.test_dir.mkdir(exist_ok=True)
        
        # Create sample files with national numbers
        self.sample_files = [
            "test1234567890.txt",
            "document_with_0123456789.doc",
            "987654321_report.pdf",
            "no_national_number.txt"
        ]
        
        # Create all sample files
        for file_name in self.sample_files:
            file_path = self.test_dir / file_name
            file_path.touch()
            
        # Create a subdirectory with more files
        subdir = self.test_dir / "subdir"
        subdir.mkdir(exist_ok=True)
        for file_name in ["subdir_1234567890.txt", "subdir_no_number.txt"]:
            file_path = subdir / file_name
            file_path.touch()
            
        # Store the original working directory
        self.original_dir = os.getcwd()
        # Change to test directory
        os.chdir(self.test_dir)

    def tearDown(self):
        """Clean up test files and directories"""
        os.chdir(self.original_dir)
        shutil.rmtree(self.test_dir)
        
        # Remove the Excel output file if it exists
        excel_file = Path("matching_files.xlsx")
        if excel_file.exists():
            excel_file.unlink()

    def test_file_scanning(self):
        """Test if the scanner correctly identifies files with national numbers"""
        # Run the scanner
        scan_files()
        
        # Check if Excel file was created
        self.assertTrue(Path("matching_files.xlsx").exists())
        
        # Read the Excel file
        df_all = pd.read_excel("matching_files.xlsx", sheet_name="All Matches", dtype={"NationalNumber": str})
        df_unique = pd.read_excel("matching_files.xlsx", sheet_name="Unique Numbers", dtype={"NationalNumber": str})
        
        # Test cases
        self.assertEqual(len(df_all), 4)  # Should find 4 files with numbers
        self.assertEqual(len(df_unique), 3)  # Should find 4 unique numbers
        
        # Check if specific numbers are found
        expected_numbers = {
            "1234567890",
            "0123456789",
            "0987654321",  # 9-digit number with added leading zero
        }
        found_numbers = set(df_unique["NationalNumber"].astype(str))
        self.assertEqual(expected_numbers, found_numbers)

    def test_leading_zeros(self):
        """Test if leading zeros are preserved in the Excel file"""
        scan_files()
        
        df = pd.read_excel("matching_files.xlsx", sheet_name="All Matches", dtype={"NationalNumber": str})
        
        # Check if numbers with leading zeros are preserved
        numbers_with_leading_zeros = df[df["NationalNumber"].astype(str).str.startswith("0")]
        self.assertGreater(len(numbers_with_leading_zeros), 0)
        
        # Verify that leading zeros are actually preserved
        for num in numbers_with_leading_zeros["NationalNumber"]:
            self.assertEqual(len(str(num)), 10)

    def test_file_paths(self):
        """Test if file paths are correctly recorded"""
        scan_files()
        
        df = pd.read_excel("matching_files.xlsx", sheet_name="All Matches")
        
        # Check if all file paths exist
        for path in df["FilePath"]:
            self.assertTrue(Path(path).exists())
        
        # Check if subdirectory files are included
        subdir_files = df[df["FilePath"].str.contains("subdir")]
        self.assertGreater(len(subdir_files), 0)

    def test_excel_formatting(self):
        """Test if Excel file is properly formatted"""
        scan_files()
        
        # Use openpyxl to check Excel formatting
        import openpyxl
        wb = openpyxl.load_workbook("matching_files.xlsx")
        
        # Check if both sheets exist
        self.assertIn("All Matches", wb.sheetnames)
        self.assertIn("Unique Numbers", wb.sheetnames)
        
        # Check if tables are present
        sheet = wb["All Matches"]
        self.assertTrue(sheet.tables)  # Should have at least one table
        
        # Check if first column is formatted as text
        first_cell = sheet["A2"]  # First data cell
        self.assertEqual(first_cell.number_format, "@")

if __name__ == "__main__":
    unittest.main(verbosity=2)